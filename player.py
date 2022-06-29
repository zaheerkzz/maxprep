import requests
from bs4 import BeautifulSoup as Soup
from maxpreps.scrappers.settings import Settings as settings
from .states import States
from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl as op
import os
import datetime
import re


class Players:

    FILE_NAME = 'players_.xlsx'
    SEASON = '20-21/roster/'
    sheet_title = 'player_info'
    wb_keywords = ''
    ws_keywords = ''
    dc_id = 1000
    STATE = ''

    def __init__(self):
        if os.path.exists(self.FILE_NAME):
            self.wb_keywords = op.load_workbook(self.FILE_NAME)
            self.ws_keywords = self.wb_keywords.active
        else:
            self.wb_keywords = Workbook()
            self.ws_keywords = self.wb_keywords.active
            # self.ws_keywords.append(["imageUrl", "name", "school", "jersy", "position",
            #                         "player attributes", "explaination", "featured statistics", "statistics"])
            self.ws_keywords.append(["Data capture ID", "Manual-ID", "Timestamp", "Priority", "Complete",
                                     "Name Scrape", "First Name", "Last Name", "Confirmed Name", "gather 2 link", "twitter link",
                                     "Twitter DM", "twitter header", "HUDL Highlight link", "City", "State", "Jersy", "Position",
                                     "height", 'weight lbs', "GPA", "ACT", "SAT", "40 time", "Bench", "Squat", "Class", "Graduation Year"
                                     "school", "Image Link", "player attributes", "explaination", "featured statistics", "statistics"])
            self.wb_keywords.save(self.FILE_NAME)

    teamUrl = []
    detailUrls = []
    playersUrl = 'https://www.maxpreps.com/nc/brevard/brevard-blue-devils/football/' + SEASON

    def get_data_capture_id(self):
        self.dc_id = self.dc_id + 1
        return f"{(self.STATE[:2]).upper()}{self.dc_id}"

    def get_manual_id(self):
        return f"{(self.STATE[:2]).upper()}{'00'}{self.dc_id}"

    def get_state_teams(self, state_name):
        self.STATE = state_name
        # excel fileName to save
        # self.FILE_NAME = self.FILE_NAME + state_name.lower().replace(' ', '_') + \
        #     '_' + self.SEASON

        counter = 1
        state_url = States.get_state_url(state_name=state_name)
        response = requests.get(url=state_url, headers=settings.headers)
        pageSoup = Soup(response.text, 'html.parser')
        t_length = len(pageSoup.findAll(
            'div', class_='StyledLinksCard__StyledLinksGrid-sc-1l6hysw-1 hEdUIf'))
        print(t_length)
        teams = pageSoup.findAll(
            'div', class_='StyledLinksCard__StyledLinksGrid-sc-1l6hysw-1 hEdUIf')[t_length-1]
        for team in teams.findAll('a'):
            # if counter > 64:
            urlresp = requests.get(url=team['href'], headers=settings.headers)
            print(str(counter) + " : " + urlresp.url + self.SEASON)
            self.get_players(urlresp.url + self.SEASON)
            counter = counter + 1
            # else:
            #     counter = counter + 1

    def get_players(self, team_url):
        response = requests.get(url=team_url, headers=settings.headers)
        pageSoup = Soup(response.text, 'html.parser')
        players = pageSoup.find('table')
        if players is not None:
            for player in players.findAll('tr'):
                detailUrl = player.find('a')['href']
                detailUrl = detailUrl.replace(
                    'default.htm', 'football/stats.htm')
                player_attrinutes = player.findAll('td')[2].find('div').text
                if 'SS' in player_attrinutes:
                    self.player_details(detailUrl)

    def player_details(self, player_url):
        maxPrepDict = {}
        # url = 'https://www.maxpreps.com/athlete/kylei-richmond/HPq9Pq4uEemA0ZFtwg1OMQ/football/stats.htm'
        response = requests.get(url=player_url, headers=settings.headers)
        pageSoup = Soup(response.text, 'html.parser')

        # player information
        playerInfo = pageSoup.find('div', class_='athlete-info')
        if playerInfo is not None:
            try:
                playerImage = playerInfo.find(
                    'div', class_='athlete-photo').find('a')['style']
                playerImage = playerImage.split(
                    'background-image:url("')[1].split('.jpg')[0] + '.jpg' if playerImage is not None else ""
            except:
                playerImage = "not-found"
            maxPrepDict['imageUrl'] = playerImage
            # name and school
            playerNames = playerInfo.find(
                'div', class_='athlete-name-school-name')
            if playerNames is not None:
                playerName = playerNames.find('h1', class_='athlete-name').a
                playerName = playerName.text if playerName is not None else ""
                schoolName = playerNames.find('div', class_='row').a
                schoolName = schoolName.text if schoolName is not None else ""

                maxPrepDict['name'] = playerName
                maxPrepDict['school'] = schoolName

        # season info
        seasonInfo = pageSoup.find('div', class_='season-info')
        if seasonInfo is not None:
            jersyPos = seasonInfo.find('div', class_='row').find('dl')
            if jersyPos is not None:
                jersyPos = jersyPos.find('dd')
                jpLen = len(jersyPos.findAll('span'))
                if jpLen == 2:
                    jersy = jersyPos.findAll('span')[0]
                    pos = jersyPos.findAll('span')[1]
                    if jersy is not None:
                        maxPrepDict['jersy'] = jersy.text
                    if pos is not None:
                        maxPrepDict['position'] = pos.text
                elif jpLen == 1:
                    jersy = jersyPos.find('span')
                    if jersy is not None:
                        maxPrepDict['jersy'] = jersy.text
                    maxPrepDict['position'] = ""
                else:
                    maxPrepDict['jersy'] = ""
                    maxPrepDict['position'] = ""

        # player attributes
            playerAttributes = playerInfo.find(
                'div', class_='athlete-attributes')
            if playerAttributes is not None:
                playerAttr = {}
                height = playerAttributes.find('span', class_='height')
                height = height.text if height is not None else ""

                weight = playerAttributes.find('span', class_='weight')
                weight = weight.text if weight is not None else ""

                grade = playerAttributes.find('span', class_='grade')
                grade = grade.text if grade is not None else ""

                graduationYear = playerAttributes.find(
                    'span', class_='graduation-year')
                graduationYear = graduationYear.text if graduationYear is not None else ""

                playerAttr['height'] = height.replace("'", '').replace('"', '')
                playerAttr['weight'] = weight
                playerAttr['grade'] = grade
                playerAttr['graduation Year'] = graduationYear

                maxPrepDict['player attributes'] = playerAttr

        # player details and statistics
        playerDetails = pageSoup.find('div', class_='content-center')
        if playerDetails is not None:
            explaination = playerDetails.find('p', class_='explanation')
            explaination = explaination.text if explaination is not None else ""
            maxPrepDict['explaination'] = explaination

            # featured statistics
            faeturedStats = playerDetails.find('ul', class_='featured-stats')
            if faeturedStats is not None:
                featured = {}
                for li in faeturedStats.findAll('li'):
                    if li is not None:
                        stateName = li.find('div', class_='stat-name')
                        stateField = li.find('div', class_='stat-field')
                        featured[stateName.text.strip(
                        )] = stateField.text.strip()
                maxPrepDict['featured statistics'] = featured

        # player statistics and detail information
        if playerDetails is not None:
            playerStats = playerDetails.find('div', class_='stats-grids')
            # player statistics
            if playerStats is not None:
                statsDict = {}
                for stats in playerStats.findAll('div'):
                    if stats is not None:
                        statsHeader = stats.h3
                        if statsHeader is not None:
                            # offense sport
                            if statsHeader.text == "Offense":
                                offenseDict = {}
                                # print(statsHeader.text)
                                for statDiv in stats.findAll('div'):
                                    statDivHeader = statDiv.h4
                                    if statDivHeader is not None:
                                        # print(statDivHeader.text)
                                        statTable = statDiv.table
                                        offenseDict[statDivHeader.text] = self.parse_table(
                                            statTable)
                                statsDict['offense'] = offenseDict

                            # defense sport
                            elif statsHeader.text == "Defense":
                                defenseDict = {}
                                # print(statsHeader.text)
                                for statDiv in stats.findAll('div'):
                                    statDivHeader = statDiv.h4
                                    if statDivHeader is not None:
                                        # print(statDivHeader.text)
                                        statTable = statDiv.table
                                        defenseDict[statDivHeader.text] = self.parse_table(
                                            statTable)
                                statsDict['defense'] = defenseDict

                            # player scoring stats
                            elif statsHeader.text == "Scoring":
                                scoringDict = {}
                                # print(statsHeader.text)
                                for statDiv in stats.findAll('div'):
                                    statDivHeader = statDiv.h4
                                    if statDivHeader is not None:
                                        # print(statDivHeader.text)
                                        statTable = statDiv.table
                                        scoringDict[statDivHeader.text] = self.parse_table(
                                            statTable)
                                statsDict['scoring'] = scoringDict
                maxPrepDict['statistics'] = statsDict
        # print(str(maxPrepDict).replace("'", '"'))
        self.write_to_csv(maxPrepDict)

    def parse_table(self, statTable):
        dictList = []
        if statTable is not None:
            tableDict = {}
            thead = statTable.thead.tr
            # table head data
            tHead = []
            for td in thead.findAll('th'):
                tHead.append(td.text)
            # table body data
            tBody = []
            for tbody in statTable.tbody.findAll('tr'):
                tData = []
                for td in tbody.findAll('td'):
                    if td is not None:
                        tData.append(td.text)
                    elif td is None or td == "":
                        tData.append('-')
                tBody.append(tData)

            for i in range(len(tBody)):
                for j in range(len(tHead)):
                    tableDict[tHead[j]] = tBody[i][j]
                dictList.append(tableDict)
            return dictList

    def write_to_csv(self, player_dict):
        # save data to excel
        try:
            self.ws_keywords.append([
                self.get_data_capture_id(), self.get_manual_id(),
                str(datetime.datetime.now()), "", "",
                player_dict['name'], self.get_name(player_dict['name'])[0],
                self.get_name(player_dict['name'])[1], "", "", "",
                "", "", "", "City", self.STATE,
                player_dict['jersy'],
                "Position",
                player_dict['player attributes']['height'],
                player_dict['player attributes']['weight'],
                "GPA", "ACT", "SAT", "40 time", "Bench", "Squat", "Class",
                player_dict['player attributes']['graduation Year'],
                player_dict['school'], player_dict['imageUrl'],
                str(player_dict['player attributes']),
                str(player_dict['explaination']),
                str(player_dict['featured statistics']),
                str(player_dict['statistics'])
            ])
            # self.ws_keywords.append([player_dict['imageUrl'], player_dict['name'], player_dict['school'], player_dict['jersy'], player_dict['position'], str(
            #     player_dict['player attributes']), str(player_dict['explaination']), str(player_dict['featured statistics']), str(player_dict['statistics'])])
            self.wb_keywords.save(self.FILE_NAME)
        except:
            pass
        # df = pd.DataFrame(player_dict).T  # transpose to look just like the sheet above
        # df.to_csv('fplayers_data.csv', mode='a')

    def get_names(self, player_name):
        return re.findall('[A-Z][a-z]*', player_name)
